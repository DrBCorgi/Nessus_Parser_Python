import argparse
import os
import datetime
from lxml import etree
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

# Global data structures
host_scan_data = []
port_data = set()

def normalize_host_data(host, recast_severity=False):
    host_name = host.find('HostProperties/tag[@name="host-ip"]').text
    host_fqdn_elem = host.find('HostProperties/tag[@name="host-fqdn"]')
    host_fqdn = host_fqdn_elem.text if host_fqdn_elem is not None else ''
    report_items = host.findall('ReportItem') if host.find('ReportItem') is not None else []

    for item in report_items:
        severity = int(item.get('severity'))
        severity_map = {0: 'Info', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical'}
        severity_text = severity_map.get(severity, 'Info')

        port = item.get('port', '0')
        protocol = item.get('protocol', '')
        svc_name = item.get('svc_name', '')
        plugin_id = item.get('pluginID', '')
        plugin_name = item.get('pluginName', '')
        synopsis = item.findtext('synopsis', 'N/A')
        description = item.findtext('description', 'N/A')
        solution = item.findtext('solution', 'N/A')
        cvss_base_score = item.findtext('cvss_base_score', 'N/A')
        cvss_vector = item.findtext('cvss_vector', 'N/A')
        plugin_output = item.findtext('plugin_output', 'N/A')
        cve = ', '.join([c.text for c in item.findall('cve')]) if item.findall('cve') else 'N/A'

        vuln_line = f"{host_name}\t{host_fqdn}\t{protocol}\t{port}\t{severity_text}\t{plugin_name}\t{synopsis}\t{description}\t{solution}\t{cvss_base_score}\t{cvss_vector}\t{plugin_output}\t{cve}\t{plugin_id}"
        host_scan_data.append(vuln_line)

        if port != '0':
            port_key = f"{host_name}:{host_fqdn}:{port}:{protocol}:{svc_name}"
            port_data.add(port_key)

def extract_vulnerabilities(report_hosts):
    vulns = {}
    for host in report_hosts:
        host_id = host.find('HostProperties/tag[@name="host-ip"]').text
        report_items = host.findall('ReportItem') if host.find('ReportItem') is not None else []
        for item in report_items:
            plugin_id = item.get('pluginID')
            key = f"{host_id}-{plugin_id}"
            vulns[key] = {
                'host_id': host_id,
                'plugin_id': plugin_id,
                'plugin_name': item.get('pluginName'),
                'severity': item.get('severity')
            }
    return vulns

def create_styles(wb):
    styles = {}
    styles['center_border6'] = {
        'font': Font(bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='000000', end_color='000000', fill_type='solid'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['cell'] = {
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['url'] = {
        'font': Font(color='0000FF', underline='single')
    }
    styles['critical'] = {
        'font': Font(bold=True),
        'fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['high'] = {
        'font': Font(bold=True),
        'fill': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['medium'] = {
        'font': Font(bold=True),
        'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['low'] = {
        'font': Font(bold=True),
        'fill': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['info'] = {
        'font': Font(bold=True),
        'fill': PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='left', vertical='top')
    }
    styles['center_border3'] = {
        'font': Font(bold=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    styles['center_border1'] = {
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    return styles

def apply_style(cell, style_dict):
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'border' in style_dict:
        cell.border = style_dict['border']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']

def main():
    parser = argparse.ArgumentParser(description='Parse Nessus XMLv2 files and provide an Excel report.')
    parser.add_argument('-d', help='Directory containing Nessus XMLv2 scan files to parse.')
    parser.add_argument('-f', help='Single Nessus XMLv2 scan file to parse.')
    parser.add_argument('-r', action='store_true', help='Recast severity values (currently same mapping)')
    parser.add_argument('-o', default='nessus_report', help='Output prefix name for report (default: nessus_report)')
    parser.add_argument('-p', help='Specify the previous Nessus XMLv2 scan file for comparison.')
    parser.add_argument('-c', help='Specify the current Nessus XMLv2 scan file for comparison.')
    parser.add_argument('-v', action='store_true', help='Display this help message.')
    args = parser.parse_args()

    if args.v or (not any([args.d, args.f, args.p, args.c])):
        print(parser.format_help())
        return

    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    report_prefix = args.o
    report_file = f"{report_prefix}_{timestamp}.xlsx"

    if args.p and args.c:
        # Comparison mode
        previous_file = args.p
        current_file = args.c
        dir_path = os.path.dirname(current_file) or '.'
        output_path = os.path.join(dir_path, report_file)

        previous_tree = etree.parse(previous_file)
        current_tree = etree.parse(current_file)

        previous_hosts = previous_tree.findall('.//ReportHost')
        current_hosts = current_tree.findall('.//ReportHost')

        previous_vulns = extract_vulnerabilities(previous_hosts)
        current_vulns = extract_vulnerabilities(current_hosts)

        closed_out = [v for k, v in previous_vulns.items() if k not in current_vulns]
        new_vulns = [v for k, v in current_vulns.items() if k not in previous_vulns]

        wb = openpyxl.Workbook()
        styles = create_styles(wb)

        # Closed Out Worksheet
        closed_ws = wb.create_sheet('Closed Out Vulnerabilities')
        closed_ws['B1'].hyperlink = '#\'Home Worksheet\'!A1'
        closed_ws['B1'].value = 'Home'
        apply_style(closed_ws['B1'], styles['url'])
        headers = ['Host Identifier', 'Plugin ID', 'Plugin Name', 'Severity']
        for col, header in enumerate(headers, 1):
            cell = closed_ws.cell(row=2, column=col, value=header)
            apply_style(cell, styles['center_border6'])
        closed_ws.freeze_panes = 'C3'
        closed_ws.auto_filter.ref = 'A2:D2'
        for col in range(1, 5):
            closed_ws.column_dimensions[get_column_letter(col)].width = 20
        for row, vuln in enumerate(closed_out, 3):
            closed_ws.cell(row=row, column=1, value=vuln['host_id'])
            closed_ws.cell(row=row, column=2, value=vuln['plugin_id'])
            closed_ws.cell(row=row, column=3, value=vuln['plugin_name'])
            closed_ws.cell(row=row, column=4, value=vuln['severity'])
            for col in range(1, 5):
                apply_style(closed_ws.cell(row=row, column=col), styles['cell'])

        # New Vulnerabilities Worksheet
        new_ws = wb.create_sheet('New Vulnerabilities')
        new_ws['B1'].hyperlink = '#\'Home Worksheet\'!A1'
        new_ws['B1'].value = 'Home'
        apply_style(new_ws['B1'], styles['url'])
        for col, header in enumerate(headers, 1):
            cell = new_ws.cell(row=2, column=col, value=header)
            apply_style(cell, styles['center_border6'])
        new_ws.freeze_panes = 'C3'
        new_ws.auto_filter.ref = 'A2:D2'
        for col in range(1, 5):
            new_ws.column_dimensions[get_column_letter(col)].width = 20
        for row, vuln in enumerate(new_vulns, 3):
            new_ws.cell(row=row, column=1, value=vuln['host_id'])
            new_ws.cell(row=row, column=2, value=vuln['plugin_id'])
            new_ws.cell(row=row, column=3, value=vuln['plugin_name'])
            new_ws.cell(row=row, column=4, value=vuln['severity'])
            for col in range(1, 5):
                apply_style(new_ws.cell(row=row, column=col), styles['cell'])

        # Home Worksheet
        home_ws = wb.create_sheet('Home Worksheet', 0)
        home_ws['A1'] = 'Comparison Report'
        home_ws.merge_cells('A1:B1')
        apply_style(home_ws['A1'], styles['center_border6'])
        home_ws['A2'] = 'Worksheets'
        apply_style(home_ws['A2'], styles['center_border3'])
        home_ws['A3'].hyperlink = '#\'Closed Out Vulnerabilities\'!A1'
        home_ws['A3'].value = 'Closed Out Vulnerabilities'
        apply_style(home_ws['A3'], styles['url'])
        home_ws['A4'].hyperlink = '#\'New Vulnerabilities\'!A1'
        home_ws['A4'].value = 'New Vulnerabilities'
        apply_style(home_ws['A4'], styles['url'])

        wb.remove(wb['Sheet'])  # Remove default sheet
        wb.active = home_ws
        wb.save(output_path)
        print(f"\n\nCompleted comparison. The data is stored in {output_path}\n")

    elif args.d or args.f:
        # Regular reporting mode
        xml_files = []
        if args.d and args.f:
            print("Please only use a file or directory as a command line argument.\n")
            print(parser.format_help())
            return
        elif args.d:
            dir_path = args.d
            xml_files = [os.path.join(dir_path, f) for f in os.listdir(dir_path) if f.endswith('.nessus')]
            print(f"\nFound {len(xml_files)} .nessus files in {dir_path}\n")
        elif args.f:
            xml_files = [args.f]
            dir_path = os.path.dirname(args.f) or '.'
            print(f"\nFound {len(xml_files)} .nessus files from command line argument\n")

        for file in xml_files:
            print(f"---------  Parsing {file}\n")
            tree = etree.parse(file)
            report_hosts = tree.findall('.//ReportHost')
            for host in report_hosts:
                normalize_host_data(host, args.r)

        print("\nCompleted parsing Nessus XML data\n")
        output_path = os.path.join(dir_path, report_file)
        wb = openpyxl.Workbook()
        styles = create_styles(wb)

        # Create worksheets
        worksheets = {
            'Critical': wb.create_sheet('Critical'),
            'High': wb.create_sheet('High'),
            'Medium': wb.create_sheet('Medium'),
            'Low': wb.create_sheet('Low'),
            'Info': wb.create_sheet('Info'),
            'Port Summary': wb.create_sheet('Port Summary'),
            'Home Worksheet': wb.create_sheet('Home Worksheet', 0),
            'Summary Worksheet': wb.create_sheet('Summary Worksheet'),
            'Port Summary Worksheet': wb.create_sheet('Port Summary Worksheet')
        }

        # Headers for vulnerability worksheets
        vuln_headers = ['IP Address', 'Host Name', 'Protocol', 'Port', 'Name', 'Synopsis', 'Description', 'Solution', 'CVSS Base Score', 'CVSS Vector', 'Plugin Output', 'CVE', 'Plugin ID']
        port_headers = ['IP Address', 'Host Name', 'Port', 'Protocol', 'Service']

        for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
            ws = worksheets[severity]
            ws['B1'].hyperlink = '#\'Home Worksheet\'!A1'
            ws['B1'].value = 'Home'
            apply_style(ws['B1'], styles['url'])
            for col, header in enumerate(vuln_headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                apply_style(cell, styles['center_border6'])
            ws.freeze_panes = 'C3'
            ws.auto_filter.ref = 'A2:M2'
            for col in range(1, 14):
                ws.column_dimensions[get_column_letter(col)].width = 20 if col <= 4 else 40 if col <= 11 else 20

        # Port Summary Worksheet
        port_ws = worksheets['Port Summary']
        port_ws['B1'].hyperlink = '#\'Home Worksheet\'!A1'
        port_ws['B1'].value = 'Home'
        apply_style(port_ws['B1'], styles['url'])
        for col, header in enumerate(port_headers, 1):
            cell = port_ws.cell(row=2, column=col, value=header)
            apply_style(cell, styles['center_border6'])
        port_ws.freeze_panes = 'C3'
        port_ws.auto_filter.ref = 'A2:E2'
        for col in range(1, 6):
            port_ws.column_dimensions[get_column_letter(col)].width = 20

        # Populate vulnerability worksheets
        row_counts = {'Critical': 2, 'High': 2, 'Medium': 2, 'Low': 2, 'Info': 2}
        for line in host_scan_data:
            fields = line.strip().split('\t')
            # Ensure there are enough fields to avoid an IndexError
            if len(fields) > 4:
                severity = fields[4]
                if severity in worksheets:
                    ws = worksheets[severity]
                    row = row_counts[severity] + 1
                    # Pad fields with 'N/A' if the split resulted in fewer columns than expected
                    full_fields = fields + ['N/A'] * (len(vuln_headers) - len(fields))
                    for col, value in enumerate(full_fields[:len(vuln_headers)], 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        apply_style(cell, styles[severity.lower()])
                    row_counts[severity] += 1

        # Populate Port Summary
        port_row = 3
        for key in sorted(port_data):
            name, fqdn, port, protocol, svc_name = key.split(':')
            port_ws.cell(row=port_row, column=1, value=name)
            port_ws.cell(row=port_row, column=2, value=fqdn)
            port_ws.cell(row=port_row, column=3, value=port)
            port_ws.cell(row=port_row, column=4, value=protocol)
            port_ws.cell(row=port_row, column=5, value=svc_name)
            for col in range(1, 6):
                apply_style(port_ws.cell(row=port_row, column=col), styles['cell'])
            port_row += 1

        # Summary Worksheet
        summary_ws = worksheets['Summary Worksheet']
        summary_ws['B1'].hyperlink = '#\'Home Worksheet\'!A1'
        summary_ws['B1'].value = 'Home'
        apply_style(summary_ws['B1'], styles['url'])
        summary_ws['A1'] = 'Summary Worksheet'
        summary_ws.merge_cells('A1:D1')
        apply_style(summary_ws['A1'], styles['center_border6'])
        summary_headers = ['Critical', 'High', 'Medium', 'Low', 'Info', 'Total']
        for col, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=2, column=col, value=header)
            apply_style(cell, styles['center_border3'])
        totals = {k: v - 2 for k, v in row_counts.items()}
        total_vuln = sum(totals.values())
        for col, value in enumerate([totals['Critical'], totals['High'], totals['Medium'], totals['Low'], totals['Info'], total_vuln], 1):
            cell = summary_ws.cell(row=3, column=col, value=value)
            apply_style(cell, styles['center_border1'])
        for col in range(1, 7):
            summary_ws.column_dimensions[get_column_letter(col)].width = 20

        # Pie Chart
        chart = PieChart()
        labels = Reference(summary_ws, min_col=1, min_row=2, max_col=5, max_row=2)
        data = Reference(summary_ws, min_col=1, min_row=3, max_col=5, max_row=3)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = 'Vulnerability Distribution'
        summary_ws.add_chart(chart, 'A5')

        # Port Summary Worksheet (sorted)
        port_summary_ws = worksheets['Port Summary Worksheet']
        port_summary_ws['B1'].hyperlink = '#\'Home Worksheet\'!A1'
        port_summary_ws['B1'].value = 'Home'
        apply_style(port_summary_ws['B1'], styles['url'])
        port_summary_ws['A1'] = 'Port Summary Worksheet'
        port_summary_ws.merge_cells('A1:E1')
        apply_style(port_summary_ws['A1'], styles['center_border6'])
        for col, header in enumerate(port_headers, 1):
            cell = port_summary_ws.cell(row=2, column=col, value=header)
            apply_style(cell, styles['center_border6'])
        port_summary_ws.freeze_panes = 'C3'
        port_summary_ws.auto_filter.ref = 'A2:E2'
        for col in range(1, 6):
            port_summary_ws.column_dimensions[get_column_letter(col)].width = 20
        sorted_ports = sorted([tuple(k.split(':')) for k in port_data], key=lambda x: (x[0], int(x[2]), x[3]))
        for row, (name, fqdn, port, protocol, svc_name) in enumerate(sorted_ports, 3):
            port_summary_ws.cell(row=row, column=1, value=name)
            port_summary_ws.cell(row=row, column=2, value=fqdn)
            port_summary_ws.cell(row=row, column=3, value=port)
            port_summary_ws.cell(row=row, column=4, value=protocol)
            port_summary_ws.cell(row=row, column=5, value=svc_name)
            for col in range(1, 6):
                apply_style(port_summary_ws.cell(row=row, column=col), styles['cell'])

        # Home Worksheet
        home_ws = worksheets['Home Worksheet']
        home_ws['A1'] = 'Home Worksheet'
        home_ws.merge_cells('A1:D1')
        apply_style(home_ws['A1'], styles['center_border6'])
        home_ws['A2'] = 'Worksheets'
        apply_style(home_ws['A2'], styles['center_border3'])
        home_ws['B2'] = 'Total'
        apply_style(home_ws['B2'], styles['center_border3'])
        row = 3
        for ws_name, total in [('Critical', totals['Critical']), ('High', totals['High']), ('Medium', totals['Medium']), ('Low', totals['Low']), ('Info', totals['Info'])]:
            home_ws.cell(row=row, column=1).hyperlink = f'#\'{ws_name}\'!A1'
            home_ws.cell(row=row, column=1).value = f'{ws_name} Vulnerabilities'
            apply_style(home_ws.cell(row=row, column=1), styles['url'])
            home_ws.cell(row=row, column=2, value=total)
            apply_style(home_ws.cell(row=row, column=2), styles['center_border1'])
            row += 1
        home_ws.cell(row=row, column=1).hyperlink = '#\'Port Summary\'!A1'
        home_ws.cell(row=row, column=1).value = 'Port Summary'
        apply_style(home_ws.cell(row=row, column=1), styles['url'])
        row += 1
        home_ws.cell(row=row, column=1).hyperlink = '#\'Summary Worksheet\'!A1'
        home_ws.cell(row=row, column=1).value = 'Summary Worksheet'
        apply_style(home_ws.cell(row=row, column=1), styles['url'])
        row += 1
        home_ws.cell(row=row, column=1).hyperlink = '#\'Port Summary Worksheet\'!A1'
        home_ws.cell(row=row, column=1).value = 'Port Summary Worksheet'
        apply_style(home_ws.cell(row=row, column=1), styles['url'])
        for col in range(1, 5):
            home_ws.column_dimensions[get_column_letter(col)].width = 30

        wb.remove(wb['Sheet'])  # Remove default sheet
        wb.active = home_ws
        wb.save(output_path)
        print(f"\n\nCompleted parsing and sorting data.\nThe data is stored in {output_path}\n")

if __name__ == '__main__':
    main()